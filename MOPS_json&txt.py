#!/usr/bin/env python

# coding: utf-8

import time

import json

from openpyxl import Workbook, load_workbook

from openpyxl.utils import get_column_letter

from selenium import webdriver

from selenium.webdriver.common.keys import Keys

from selenium.webdriver.common.by import By

from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.support import expected_conditions as ec


#'89','90','91','92','93','94','95','96','97','98','99','100','101','102','103','104','105','106','107'
def init():

    year_list = ['91']#

    twse_list = ['2881'] #,'2885','2886','2887'

    tpex_list = ['4102','4103','4104','4105','4106','4107','4108','4109','4110','4111','4113','4114','4205']



    #print(len(year_list))

    #print(len(tpex_list))

    

    return year_list, twse_list, tpex_list



def driver_open():

    browser = webdriver.Chrome()

    browser.get('http://mops.twse.com.tw/mops/web/t05st01')

    return browser



def driver_close(browser):

    browser.quit()



def input_text(index, xpath):

    inputbox = browser.find_element_by_xpath(xpath)

    inputbox.clear()

    inputbox.send_keys(str(index) + Keys.RETURN)



def WebWaitXpath(s):

    try:

        xpath = s

        WebDriverWait(browser, 2).until(ec.visibility_of_element_located((By.XPATH, xpath)))

        return True

    except Exception as e:

        print(e)



def ChangeToPopUpWindow(index):

    browser.find_element_by_xpath('//*[@id="t05st01_fm"]/table/tbody/tr[' + str(index) + ']/td[6]/input').click()

    window_after = browser.window_handles[1] #獲取彈出視窗資訊

    browser.switch_to_window(window_after) #焦點切換到彈出視窗

def BackToSourceWindow(window_before):

    browser.close() #關閉彈出視窗

    browser.switch_to_window(window_before) #將焦點切回原先視窗

def ListToDict(length, l_title, l_content):

    d_details = {}

    for i in range(0, length):

        d_details[l_title[i]] = l_content[i]

    return d_details



def get_data():

    try:

        time.sleep(10) #等待5s 再次點擊

        WebWaitXpath('//*[@id="table01"]/table[2]/tbody/tr[1]/td/b') #等待元件讀取

        l_title = []

        l_content = []

        table_path = '//*[@id="table01"]/table[3]/tbody'

        col_path = browser.find_elements_by_xpath(table_path + '/tr') #欄位置

        for i in range(1, len(col_path) + 1):

            row_path = browser.find_elements_by_xpath(table_path + '/tr[' + str(i) + ']/td') #列位置

            for j in range(1, len(row_path) + 1):

                if not (j % 2 == 0):

                    title = browser.find_element_by_xpath(table_path + '/tr[' + str(i) + ']/td[' + str(j) + ']').text #標題

                    l_title.append(title)

                elif(i == len(col_path) and j == len(row_path)):

                    content = browser.find_element_by_xpath(table_path + '/tr[' + str(i) + ']/td[' + str(j) + ']').text.split('\n') #說明部分切割成List

                    for k in content:

                        k.lstrip().rstrip()

                    l_content.append(content)

                else:

                    content = browser.find_element_by_xpath(table_path + '/tr[' + str(i) + ']/td[' + str(j) + ']').text.lstrip().rstrip() #內容 去左右空白

                    l_content.append(content)

        return ListToDict(len(l_title), l_title, l_content)

    except:

        print('Get Data Error!')

        return False



def input_data(sheet, d_details):

    data = [['序號', d_details['序號']],

        ['發言日期', d_details['發言日期']],

        ['發言時間', d_details['發言時間']],

        ['發言人', d_details['發言人']],

        ['發言人職稱', d_details['發言人職稱']],

        ['發言人電話', d_details['發言人電話']],

        ['主旨', d_details['主旨']],

        ['符合條款', d_details['符合條款']],

        ['事實發生日', d_details['事實發生日']]]

    for row in data:

        sheet.append(row) #輸入資料

    for i in range(0, len(d_details['說明'])):

        if(i == 0):

            sheet.append(['說明', d_details['說明'][i]])

        else:

            sheet.append(['', d_details['說明'][i]])


def input_data2(Listed_id, Listed_year,d_details,predate,x,k,btn_details):#json,txt
    
    
    _filename = str(Listed_id + '_' + Listed_year )
    sheet= {}
    colarray = []
    arraydata={}
    arraydata["序號"]=d_details['序號']
    arraydata["發言日期"]=d_details['發言日期']
    arraydata["發言時間"]=d_details['發言時間']
    arraydata["發言人"]= d_details['發言人']
    arraydata["發言人職稱"]=d_details['發言人職稱']
    arraydata["發言人電話"]=d_details['發言人電話']
    arraydata["主旨"]=d_details['主旨']
    arraydata["符合條款"]=d_details['發言人職稱']
    arraydata["事實發生日"]=d_details['事實發生日']
    arraydata["說明"]=d_details['說明']
    
    colarray.append(arraydata)
    sheet[_filename]=colarray
    data = json.dumps(sheet,ensure_ascii=False)
    
    date=str(Listed_id) + '_' + d_details['發言日期'].replace('/', '')
    print(predate)
    print(d_details['發言日期'])
    
    with open (_filename+'.json','a+',encoding='utf-8') as f:
        if (k == 2):
            f.write('['+str(data)+",")
        elif(k==btn_details):
            f.write(str(data)+']')
        else:
            f.write(str(data)+",")
        f.close()  

    
    if (predate != (d_details['發言日期'])):
        
        with open (date+'.txt','a+',encoding='utf-8') as f:
            for i in range(0,len(d_details['說明'])):
                f.write(str(d_details['說明'][i])+"\n")
            f.close()
    else:
        
        with open (date+'_'+str(d_details['序號'])+'.txt','a+',encoding='utf-8') as f:
            for i in range(0,len(d_details['說明'])):
                f.write(str(d_details['說明'][i])+"\n")
            f.close()
    predate=d_details['發言日期']
    return predate
        


def CreateExcel(Listed_id, Listed_year, d_details):

    wb = Workbook() #創建第一個工作表

    frist_sheet = wb.active

    datetime = d_details['發言日期'].replace('/', '.')

    sheet_name = datetime + '-' + d_details['序號'] #工作表名稱

    frist_sheet.title = sheet_name

    input_data(frist_sheet, d_details) #資料輸入資料表

    _filename = str(Listed_id + '-' + Listed_year + '.xlsx')

    wb.save(filename = _filename)

    print('excel_name: ', _filename)

    print('sheet_name: ', sheet_name)

    return _filename



def ReadExcel(d_details, excel_name):

    wb = load_workbook(excel_name) #讀取工作表

    datetime = d_details['發言日期'].replace('/', '.')

    sheet_name = datetime + '-' + d_details['序號']

    sheet = wb.create_sheet(sheet_name)

    input_data(sheet, d_details)

    wb.save(filename = excel_name)

    print('excel_name: ', excel_name)

    print('sheet_name: ', sheet_name)


def DataToExcel(isFrist, Listed_id, Listed_year, d_details, excel_name):

    if(isFrist): #建立工作表

        return CreateExcel(Listed_id, Listed_year, d_details)

    else: #讀取工作表並新增工作表

        ReadExcel(d_details, excel_name)



def get_year_message(Listed):

    for i in Listed:

        input_text(i, '//*[@id="co_id"]') #公司代號或簡稱

        for j in year_range_list:

            input_text(j, '//*[@id="year"]') #年度

            print('id: %s\tyear: %s' % (i, j))

            btn_search = browser.find_element_by_xpath("//input[@type='button' and @value=' 查詢 ']") #查詢按鈕

            btn_search.click()

            time.sleep(3) #等待3s

            again = True
            
            predate=''

            while(again):

                if(WebWaitXpath('//*[@id="t05st01_fm"]/table/tbody/tr[2]/td[3]')): #等待元件讀取

                    again = False

                    isFrist = True

                    excel_name = ''

                    window_before = browser.window_handles[0] #獲取來源網頁資訊

                    btn_details = browser.find_elements_by_xpath('//*[@id="t05st01_fm"]/table/tbody/tr') #詳細資料按鈕

                    for k in range(635, len(btn_details) + 1): #迭代每則重大消息按鈕

                        print('第' + str(k - 1) + '個按鈕')

                        again_data = True
                        
                        
                        
                        x=1

                        while(again_data):

                            ChangeToPopUpWindow(k) #改變視窗焦點

                            if(get_data() == False):

                                BackToSourceWindow(window_before)

                            else:

                                print('Get Data OK!')

                                again_data = False

                                d_details = get_data()
                                
                                

                        if(k == 2): #判斷是否為首個「詳細資料」按鈕

                            excel_name = DataToExcel(isFrist, i, j, d_details, excel_name)

                            isFrist = False

                        else:

                            DataToExcel(isFrist, i, j, d_details, excel_name)
                        predate=input_data2(i,j,d_details,predate,x,k,len(btn_details))
                    
                        BackToSourceWindow(window_before)

                        time.sleep(2) #等待2s 再次搜尋下一年
                    
                    

                else:

                    if (browser.find_elements_by_xpath('//*[@id="table01"]/center/h3')):

                        print('該 %s 公開發行公司不繼續公開發行！' % i)

                        break

                    else:

                        time.sleep(10) #等待10s

                        browser.refresh() #刷新網頁


if __name__ == '__main__':

    try:

        browser = driver_open()

        year_range_list, stock_Id_TWSE_Listed, stock_Id_TPEx_Listed = init()

        get_year_message(stock_Id_TWSE_Listed) #上市公司

        #get_year_message(stock_Id_TPEx_Listed) #上櫃公司

    except:
        
        driver_close(browser)

